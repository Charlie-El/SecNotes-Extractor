from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


translate_segment_member_cn = None

UOM_CN = {
    "USD": "美元",
    "EUR": "欧元",
    "CAD": "加元",
    "GBP": "英镑",
    "CNY": "人民币",
    "JPY": "日元",
    "KRW": "韩元",
    "TWD": "新台币",
    "HKD": "港元",
    "AUD": "澳元",
    "BRL": "巴西雷亚尔",
    "MXN": "墨西哥比索",
    "INR": "印度卢比",
    "SEK": "瑞典克朗",
    "DKK": "丹麦克朗",
    "ZAR": "南非兰特",
}

TABLE1_FIELDS = [
    ("CUSIP", "证券代码（美股主键）"),
    ("stock_code", "证券代码（主键）"),
    ("company_name", "公司名称"),
    ("report_period", "报告期（如 2024-12-31）"),
    ("report_type", "年报/半年报/季报"),
    ("industry_code_l1", "一级行业"),
    ("industry_code_l2", "二级行业"),
    ("industry_code_l3", "三级行业"),
    ("industry_code_l4", "四级行业"),
    ("main_business_desc", "主营业务描述原文（去噪后）"),
    ("total_revenue", "营业总收入（元）"),
    ("total_revenue_uom", "营业总收入计量单位"),
    ("total_revenue_uom_cn", "营业总收入计量单位中文翻译"),
    ("segment_disclosure_flag", "是否有业务分部披露"),
    ("segment_count", "业务分部数量"),
    ("primary_disclosure_dim", "主要披露维度"),
    ("disclosure_dims_available", "所有可用维度列表"),
    ("has_revenue_breakdown", "是否有任何形式的营收分解"),
    ("has_margin_breakdown", "是否有毛利/毛利率分解"),
]

TABLE2_FIELDS = [
    ("segment_id", "主键（自增）"),
    ("stock_code", "外键"),
    ("report_period", "外键"),
    ("segment_name", "分部名称"),
    ("segment_type", "按业务/产品/地区等划分"),
    ("segment_axis", "原始 XBRL 维度轴名称"),
    ("product_keywords", "提取的关键产品/服务词"),
    ("segment_member_country_cn", "地区中文翻译"),
    ("segment_revenue", "分部营业收入"),
    ("segment_revenue_uom", "分部营业收入计量单位"),
    ("segment_revenue_uom_cn", "分部营业收入单位中文翻译"),
    ("segment_revenue_ratio", "占总营收比例"),
    ("segment_gross_profit", "分部毛利"),
    ("segment_gross_margin", "毛利率"),
    ("segment_desc", "分部业务描述"),
    ("source_section", "business/MDA/notes"),
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
DESC_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
NOTE_FILL = PatternFill(fill_type="solid", fgColor="EDEDED")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
FONT_NORMAL = Font(name="Arial", size=10)
FONT_BOLD = Font(name="Arial", size=10, bold=True)
SOURCE_WORKBOOK_NOTE = (
    "表注释：本表按标准源表格式生成；优先保留最新季报，如该公司最新年报 filing 更晚则改用年报。"
    "对季报金额优先使用单季 QTD，缺失时使用 YTD。"
)


def clean_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def normalize_ticker(value: Any) -> str:
    text = clean_text(value).upper().replace("_", "-")
    if "." in text:
        base, suffix = text.rsplit(".", 1)
        if len(suffix) <= 2 and suffix.isalpha():
            return base
    return text


def load_dimension_revisions(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    mapping: dict[str, str] = {}
    first_row = [clean_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    second_row = [clean_text(ws.cell(2, c).value) for c in range(1, ws.max_column + 1)] if ws.max_row >= 2 else []
    if first_row[:2] == ["原始XBRL维度名称", "primary_disclosure_dim/disclosure_dims_available/segment_axis"] and second_row[:2] == ["segment_axis", "segment_type"]:
        start_row = 3
        axis_col = 1
        type_col = 2
    else:
        idx = {name: i + 1 for i, name in enumerate(first_row)}
        start_row = 2
        axis_col = idx.get("segment_axis", 1)
        type_col = idx.get("segment_type", 2)
    for r in range(start_row, ws.max_row + 1):
        axis = clean_text(ws.cell(r, axis_col).value)
        new_type = clean_text(ws.cell(r, type_col).value)
        if axis and new_type:
            mapping[axis] = new_type
    wb.close()
    return mapping


def add_country_translation(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "segment_member_country_cn" not in out.columns:
        out["segment_member_country_cn"] = None
    if translate_segment_member_cn is None:
        return out
    mask = out.get("row_type").eq("segment") if "row_type" in out.columns else pd.Series(False, index=out.index)
    for idx, row in out[mask].iterrows():
        out.at[idx, "segment_member_country_cn"] = translate_segment_member_cn(row.get("segment_member_raw"), row.get("segment_type"))
    return out


def read_template_headers(template_path: Path) -> tuple[list[str], dict[str, str]]:
    wb = load_workbook(template_path, read_only=True, data_only=True)
    ws = wb["总表"]
    descriptions = {clean_text(ws.cell(3, c).value): clean_text(ws.cell(2, c).value) for c in range(1, ws.max_column + 1)}
    headers = [clean_text(ws.cell(3, c).value) for c in range(1, ws.max_column + 1)]
    wb.close()
    return headers, descriptions


def to_annual_compatible(raw: pd.DataFrame, template_path: Path) -> pd.DataFrame:
    headers, _ = read_template_headers(template_path)
    rows: list[dict[str, Any]] = []
    for _, row in raw.iterrows():
        is_segment = clean_text(row.get("row_type")) == "segment"
        total_revenue = row.get("total_revenue_qtd")
        if pd.isna(total_revenue) or clean_text(total_revenue) == "":
            total_revenue = row.get("total_revenue_ytd")
        segment_revenue = row.get("segment_revenue_qtd")
        if pd.isna(segment_revenue) or clean_text(segment_revenue) == "":
            segment_revenue = row.get("segment_revenue_ytd")
        segment_ratio = row.get("segment_revenue_ratio_qtd")
        if pd.isna(segment_ratio) or clean_text(segment_ratio) == "":
            segment_ratio = row.get("segment_revenue_ratio_ytd")
        rec = {h: None for h in headers}
        rec.update(
            {
                "row_type": row.get("row_type"),
                "stock_code": row.get("stock_code"),
                "company_name": row.get("company_name"),
                "CUSIP": row.get("CUSIP"),
                "ticker": row.get("ticker"),
                "ticker_match_method": row.get("ticker_match_method"),
                "cik": row.get("cik"),
                "company_name_filing": row.get("company_name_filing"),
                "report_type": row.get("report_type"),
                "report_period": row.get("report_period"),
                "filed_date": row.get("filed_date"),
                "accepted_datetime": row.get("accepted_datetime"),
                "filing_month_batch": row.get("filing_month_batch"),
                "adsh": row.get("adsh"),
                "instance": row.get("instance"),
                "sic": row.get("sic"),
                "industry_l1": row.get("industry_l1"),
                "industry_l2": row.get("industry_l2"),
                "industry_l3": row.get("industry_l3"),
                "industry_l4": row.get("industry_l4"),
                "fiscal_year": row.get("fiscal_year"),
                "fiscal_period": row.get("fiscal_period"),
                "main_business_desc": row.get("evidence_text") if not is_segment else None,
                "total_revenue": total_revenue,
                "total_revenue_tag": row.get("total_revenue_tag"),
                "total_revenue_version": row.get("total_revenue_version"),
                "total_revenue_uom": row.get("total_revenue_uom"),
                "gross_profit": row.get("gross_profit_qtd") if clean_text(row.get("gross_profit_qtd")) else row.get("gross_profit_ytd"),
                "operating_income": row.get("operating_income_qtd") if clean_text(row.get("operating_income_qtd")) else row.get("operating_income_ytd"),
                "net_income": row.get("net_income_qtd") if clean_text(row.get("net_income_qtd")) else row.get("net_income_ytd"),
                "assets": row.get("assets"),
                "assets_uom": row.get("assets_uom"),
                "liabilities": row.get("liabilities"),
                "liabilities_uom": row.get("liabilities_uom"),
                "equity": row.get("equity"),
                "equity_uom": row.get("equity_uom"),
                "cash_and_equivalents": row.get("cash_and_equivalents"),
                "cash_uom": row.get("cash_uom"),
                "operating_cash_flow": row.get("operating_cash_flow_ytd"),
                "capex": row.get("capex_ytd"),
                "rd_expense": row.get("rd_expense_qtd") if clean_text(row.get("rd_expense_qtd")) else row.get("rd_expense_ytd"),
                "eps_basic": row.get("eps_basic_qtd") if clean_text(row.get("eps_basic_qtd")) else row.get("eps_basic_ytd"),
                "eps_diluted": row.get("eps_diluted_qtd") if clean_text(row.get("eps_diluted_qtd")) else row.get("eps_diluted_ytd"),
                "segment_disclosure_flag": row.get("segment_disclosure_flag"),
                "segment_count": row.get("segment_count"),
                "primary_disclosure_dim": row.get("primary_disclosure_dim"),
                "disclosure_dims_available": row.get("disclosure_dims_available"),
                "has_revenue_breakdown": row.get("has_revenue_breakdown"),
                "has_margin_breakdown": False,
                "segment_id": row.get("segment_id"),
                "segment_name": row.get("segment_name"),
                "segment_type": row.get("segment_type"),
                "segment_axis": row.get("segment_axis"),
                "segment_member_raw": row.get("segment_member_raw"),
                "segment_member_country_cn": row.get("segment_member_country_cn"),
                "segment_revenue": segment_revenue,
                "segment_revenue_uom": row.get("segment_revenue_uom"),
                "segment_revenue_uom_cn": UOM_CN.get(clean_text(row.get("segment_revenue_uom")), ""),
                "segment_revenue_ratio": segment_ratio,
                "segment_desc": row.get("evidence_text") if is_segment else None,
                "source_section": "notes" if is_segment else "company",
                "source_tag": row.get("source_tag"),
                "source_version": row.get("source_version"),
                "source_group_tag": row.get("source_tag"),
                "source_dim_hash": row.get("source_dim_hash"),
                "source_segments_text": row.get("source_segments_text"),
                "source_report": row.get("source_report"),
                "source_report_shortname": row.get("source_report_shortname"),
                "source_report_longname": row.get("source_report_longname"),
                "evidence_text": row.get("evidence_text"),
                "extraction_status": row.get("extraction_status"),
                "notes_file_batch": row.get("notes_file_batch"),
            }
        )
        rows.append(rec)
    return pd.DataFrame(rows, columns=headers)


def autosize(ws) -> None:
    for col_idx, column in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column[:120]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 42)


def write_source_workbook(df: pd.DataFrame, template_path: Path, output_path: Path) -> None:
    headers, descriptions = read_template_headers(template_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "总表"
    ws.cell(1, 1, SOURCE_WORKBOOK_NOTE)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1).fill = NOTE_FILL
    ws.cell(1, 1).font = FONT_BOLD
    ws.cell(1, 1).alignment = Alignment(wrap_text=True)
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(2, col_idx, descriptions.get(header, ""))
        ws.cell(3, col_idx, header)
        ws.cell(2, col_idx).fill = DESC_FILL
        ws.cell(3, col_idx).fill = HEADER_FILL
        ws.cell(3, col_idx).font = FONT_BOLD
    for row_idx, (_, row) in enumerate(df.iterrows(), start=4):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row_idx, col_idx, row.get(header))
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{ws.cell(max(ws.max_row, 3), len(headers)).coordinate}"
    autosize(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_table(ws, note: str, fields: list[tuple[str, str]], df: pd.DataFrame) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(fields))
    note_cell = ws.cell(1, 1, note)
    note_cell.fill = NOTE_FILL
    note_cell.font = FONT_BOLD
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col_idx, (field, desc) in enumerate(fields, start=1):
        desc_cell = ws.cell(2, col_idx, desc)
        header_cell = ws.cell(3, col_idx, field)
        for cell in (desc_cell, header_cell):
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        desc_cell.fill = DESC_FILL
        desc_cell.font = FONT_NORMAL
        header_cell.fill = HEADER_FILL
        header_cell.font = FONT_BOLD
    for row_idx, (_, row) in enumerate(df.iterrows(), start=4):
        for col_idx, (field, _) in enumerate(fields, start=1):
            cell = ws.cell(row_idx, col_idx, row.get(field))
            cell.font = FONT_NORMAL
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{ws.cell(max(ws.max_row, 3), len(fields)).coordinate}"
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 24
    autosize(ws)


def write_formatted_workbook(source_df: pd.DataFrame, output_path: Path) -> None:
    companies = source_df[source_df["row_type"] == "company"].copy()
    segments = source_df[source_df["row_type"] == "segment"].copy()
    table1 = pd.DataFrame()
    for field, _ in TABLE1_FIELDS:
        if field == "industry_code_l1":
            table1[field] = companies["industry_l1"]
        elif field == "industry_code_l2":
            table1[field] = companies["industry_l2"]
        elif field == "industry_code_l3":
            table1[field] = companies["industry_l3"]
        elif field == "industry_code_l4":
            table1[field] = companies["industry_l4"]
        elif field == "total_revenue_uom_cn":
            table1[field] = companies["total_revenue_uom"].map(lambda v: UOM_CN.get(clean_text(v), ""))
        else:
            table1[field] = companies[field] if field in companies.columns else None
    table2 = pd.DataFrame()
    for field, _ in TABLE2_FIELDS:
        if field == "product_keywords":
            table2[field] = segments["segment_member_raw"]
        elif field == "segment_revenue_uom_cn":
            table2[field] = segments["segment_revenue_uom"].map(lambda v: UOM_CN.get(clean_text(v), ""))
        elif field == "source_section":
            table2[field] = "notes"
        else:
            table2[field] = segments[field] if field in segments.columns else None

    wb = Workbook()
    wb.remove(wb.active)
    write_table(wb.create_sheet("表1_公司主营业务原文"), f"本次按结果模板生成。表1输出 {len(table1)} 行。", TABLE1_FIELDS, table1)
    write_table(wb.create_sheet("表2_公司业务分部明细"), f"本次按结果模板生成。表2输出 {len(table2)} 行。", TABLE2_FIELDS, table2)
    ws = wb.create_sheet("字段映射说明")
    headers = ["范围", "PDF 字段名", "处理结果", "原表字段名", "说明"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, f"季度提取字段映射说明。表1输出 {len(table1)} 行，表2输出 {len(table2)} 行。")
    ws.cell(1, 1).fill = NOTE_FILL
    ws.cell(1, 1).font = FONT_BOLD
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(2, c, header)
        cell.fill = HEADER_FILL
        cell.font = FONT_BOLD
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row_idx = 3
    for scope, fields in (("表1_公司主营业务原文", TABLE1_FIELDS), ("表2_公司业务分部明细", TABLE2_FIELDS)):
        for field, desc in fields:
            source = field
            if field.startswith("industry_code_"):
                source = field.replace("industry_code_", "industry_")
            elif field.endswith("_uom_cn"):
                source = field.replace("_uom_cn", "_uom")
            elif field == "product_keywords":
                source = "segment_member_raw"
            status = "字段翻译" if field.endswith("_uom_cn") else "已保留"
            for c, value in enumerate([scope, field, status, source, desc], start=1):
                cell = ws.cell(row_idx, c, value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = FONT_NORMAL
            row_idx += 1
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:E{ws.max_row}"
    autosize(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
