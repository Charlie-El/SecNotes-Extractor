from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ccxd_us_reports_app.vendor.annual import extract_notes_coverage as annual_notes
from ccxd_us_reports_app.quarterly_output_utils import write_formatted_workbook


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
DESC_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
NOTE_FILL = PatternFill(fill_type="solid", fgColor="EDEDED")
FONT_BOLD = Font(name="Microsoft YaHei UI", size=10, bold=True)


def create_source_template(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "总表"

    headers = list(annual_notes.ROW_COLUMNS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    note_cell = ws.cell(1, 1, annual_notes.TABLE_NOTE)
    note_cell.fill = NOTE_FILL
    note_cell.font = FONT_BOLD
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for idx, header in enumerate(headers, start=1):
        desc_cell = ws.cell(2, idx, annual_notes.FIELD_DESCRIPTIONS.get(header, ""))
        header_cell = ws.cell(3, idx, header)
        desc_cell.fill = DESC_FILL
        header_cell.fill = HEADER_FILL
        header_cell.font = FONT_BOLD
        desc_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A4"
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 44
    ws.row_dimensions[3].height = 24
    wb.save(path)
    return path


def create_mapped_template(path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    empty = pd.DataFrame(columns=annual_notes.ROW_COLUMNS)
    write_formatted_workbook(empty, path)
    return path
