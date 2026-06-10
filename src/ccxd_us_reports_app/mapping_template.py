from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
NOTE_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
FONT = Font(name="Microsoft YaHei UI", size=10)
FONT_BOLD = Font(name="Microsoft YaHei UI", size=10, bold=True)


def _build_template(
    path: Path,
    *,
    sheet_title: str,
    mode_name: str,
    include_examples: bool,
) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    ws["A1"] = "原始XBRL维度名称"
    ws["B1"] = "primary_disclosure_dim/disclosure_dims_available/segment_axis"
    ws["A2"] = "segment_axis"
    ws["B2"] = "segment_type"
    if include_examples:
        ws["A3"] = "StatementBusinessSegmentsAxis"
        ws["B3"] = "business"
        ws["A4"] = "ProductOrServiceAxis"
        ws["B4"] = "product"
        ws["A5"] = "GeographicalAreasAxis"
        ws["B5"] = "region"
        ws["A6"] = "OperatingSegmentsAxis"
        ws["B6"] = "business"

    notes = [
        "说明",
        "前两行表头必须保持不变。",
        f"此模板仅用于{mode_name}模式。",
        "用户需要填写的是：segment_axis 对应的 segment_type。",
        "如果运行时不提供映射表，程序不会输出“修改映射表”。",
    ]
    for idx, text in enumerate(notes, start=1):
        ws[f"D{idx}"] = text

    for row in (1, 2):
        for cell in ws[row]:
            cell.fill = HEADER_FILL
            cell.font = FONT_BOLD
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(1, len(notes) + 1):
        cell = ws[f"D{row}"]
        cell.fill = NOTE_FILL
        cell.font = FONT if row != 1 else FONT_BOLD
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for column in ("A", "B", "D"):
        ws.column_dimensions[column].width = 42

    wb.save(path)
    return path


def create_quarterly_mapping_template(path: Path, *, include_examples: bool = True) -> Path:
    return _build_template(
        path,
        sheet_title="季报映射模板",
        mode_name="季报",
        include_examples=include_examples,
    )


def create_annual_mapping_template(path: Path, *, include_examples: bool = True) -> Path:
    return _build_template(
        path,
        sheet_title="年报映射模板",
        mode_name="年报",
        include_examples=include_examples,
    )
